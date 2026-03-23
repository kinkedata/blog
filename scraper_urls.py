#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
scraper_urls.py — Recorre el listado de Telcel Empresas Tendencias
y genera un Excel con URL, Título y Fecha de cada nota.
"""

import os
import re
import sys
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ openpyxl no está instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    print("✅ Selenium instalado correctamente")
except ImportError:
    print("❌ Selenium no está instalado. Ejecuta: pip install selenium")
    sys.exit(1)

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
STOP_FILE   = os.path.join(BASE_DIR, 'stop_signal.txt')
LISTING_URL = 'https://www.telcel.com/empresas/tendencias'
TOTAL_PAGES = 192
CARD_SEL    = '.comp-tendencia-item-todas'
ITEM_SEL    = '.comp-tendencia-item-todas a'
CONTAINER   = 'comp-tendencia-contenedor-todas'

MESES = {
    'ENE': '01', 'FEB': '02', 'MAR': '03', 'ABR': '04',
    'MAY': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
    'SEP': '09', 'OCT': '10', 'NOV': '11', 'DIC': '12',
}


def parse_fecha(raw):
    """Convierte 'ENE 22, 2025' → '22/01/2025'. Retorna '' si no puede parsear."""
    raw = raw.strip()
    m = re.match(r'([A-ZÁÉÍÓÚ]+)\s+(\d{1,2}),\s*(\d{4})', raw)
    if not m:
        return raw
    mes_str, dia, anio = m.group(1), m.group(2), m.group(3)
    mes_num = MESES.get(mes_str.upper(), '00')
    return f"{dia.zfill(2)}/{mes_num}/{anio}"


def get_current_items(driver):
    """Extrae URL, título y fecha de cada card visible en la página."""
    items = []
    for card in driver.find_elements(By.CSS_SELECTOR, CARD_SEL):
        try:
            a    = card.find_element(By.TAG_NAME, 'a')
            href = a.get_attribute('href') or ''
            if '/empresas/tendencias/notas/' not in href:
                continue
            url = href.replace('https://', '').replace('http://', '').rstrip('/')

            try:
                titulo = card.find_element(By.CSS_SELECTOR, '.card-info-titulo').text.strip()
            except NoSuchElementException:
                titulo = ''

            try:
                fecha_raw = card.find_element(By.CSS_SELECTOR, '.card-date').text.strip()
                fecha = parse_fecha(fecha_raw)
            except NoSuchElementException:
                fecha = ''

            items.append({'url': url, 'titulo': titulo, 'fecha': fecha})
        except NoSuchElementException:
            continue
    return items


def get_next_page_btn(driver):
    """Retorna el <a> de la siguiente página o None si es la última."""
    try:
        active  = driver.find_element(By.CSS_SELECTOR, 'a.active[data-contenido="paginador"]')
        next_li = active.find_element(By.XPATH, '../following-sibling::li[1]')
        next_a  = next_li.find_element(By.TAG_NAME, 'a')
        if next_a.get_attribute('data-contenido') != 'paginador':
            return None
        return next_a
    except NoSuchElementException:
        return None


def click_next(driver, current_first_href):
    """Clic en siguiente página y espera a que el contenido cambie."""
    next_btn = get_next_page_btn(driver)
    if next_btn is None:
        return False

    driver.execute_script("arguments[0].click();", next_btn)

    try:
        WebDriverWait(driver, 15).until(
            lambda d: (
                d.find_elements(By.CSS_SELECTOR, ITEM_SEL) and
                d.find_element(By.CSS_SELECTOR, ITEM_SEL).get_attribute('href') != current_first_href
            )
        )
    except TimeoutException:
        print("⚠️  Timeout esperando que cargue el contenido de la página.")

    return True


def save_excel(all_items, output_file, stopped):
    """Guarda los datos en un archivo Excel con formato."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Notas Blog Telcel'

    # Estilos de cabecera
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='00529B')
    hdr_align = Alignment(horizontal='center', vertical='center')

    headers = ['URL', 'Título', 'Fecha']
    col_widths = [70, 60, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    # Datos
    for row_idx, item in enumerate(all_items, start=2):
        ws.cell(row=row_idx, column=1, value=item['url'])
        ws.cell(row=row_idx, column=2, value=item['titulo'])
        ws.cell(row=row_idx, column=3, value=item['fecha'])

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(output_file)


def main():
    print("🔍 GENERADOR DE LISTA DE URLs — BLOG TELCEL TENDENCIAS")
    print("=" * 70)
    print()

    if os.path.exists(STOP_FILE):
        os.remove(STOP_FILE)

    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)')

    print("🌐 Iniciando navegador Chrome...")
    try:
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        print(f"❌ No se pudo iniciar Chrome: {e}")
        return

    timestamp   = time.strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(BASE_DIR, f'notas_blog_telcel_{timestamp}.xlsx')

    all_items = []
    page      = 1
    stopped   = False

    try:
        driver.get(LISTING_URL)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, CONTAINER))
        )
        print(f"✅ Página cargada: {LISTING_URL}")
        print()
        print("🔄 RECORRIENDO PÁGINAS...")
        print("=" * 70)
        print()

        while True:
            # Verificar señal de parada
            if os.path.exists(STOP_FILE):
                os.remove(STOP_FILE)
                print()
                print(f"⚠️ Señal de parada recibida (página {page}).")
                stopped = True
                break

            # Extraer cards de la página actual
            page_items = get_current_items(driver)
            all_items.extend(page_items)
            print(f"⏳ Página {page}/{TOTAL_PAGES} — {len(page_items)} notas  |  Total: {len(all_items)}")

            # Guardar primer href para detectar cambio
            try:
                first_href = driver.find_element(By.CSS_SELECTOR, ITEM_SEL).get_attribute('href')
            except NoSuchElementException:
                first_href = ''

            if not click_next(driver, first_href):
                print()
                print("✅ Última página alcanzada.")
                break

            page += 1

    except KeyboardInterrupt:
        print("\n⚠️ Interrumpido por el usuario.")
        stopped = True
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
    finally:
        print()
        print("🔌 Cerrando navegador...")
        driver.quit()

    # Guardar Excel
    if all_items:
        print()
        print(f"💾 Generando Excel con {len(all_items)} notas...")
        save_excel(all_items, output_file, stopped)
        status = "PARCIAL" if stopped else "COMPLETO"
        print(f"✅ Archivo guardado ({status}): {output_file}")
        print()
        print("📊 RESUMEN:")
        print(f"   Páginas recorridas : {page} de {TOTAL_PAGES}")
        print(f"   Notas encontradas  : {len(all_items)}")
        print()
        print("✅ ¡PROCESO COMPLETADO!")
    else:
        print("❌ No se encontraron notas.")


if __name__ == '__main__':
    main()
