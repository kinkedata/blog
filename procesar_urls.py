import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── 1. Leer el CSV de visitas ──────────────────────────────────────────────────
csv_path = "Freeform table - URL sin parametros.csv"

# Las primeras 11 líneas son metadatos; la línea 12 es el header (",Visits")
df = pd.read_csv(csv_path, skiprows=11, header=0)
df.columns = ["URL", "Visits"]

# Quitar la fila de totales (primera fila contiene "URL sin parametros")
df = df[df["URL"] != "URL sin parametros"].copy()

# Limpiar: eliminar filas con URL vacía o NaN
df = df.dropna(subset=["URL"])
df["URL"] = df["URL"].str.strip()
df = df[df["URL"] != ""]

# Convertir Visits a numérico
df["Visits"] = pd.to_numeric(df["Visits"], errors="coerce").fillna(0).astype(int)

print(f"Total filas antes de consolidar: {len(df)}")
print(f"Filas con .html: {df['URL'].str.endswith('.html').sum()}")

# ── 2. Normalizar URLs quitando .html y sumar visitas ────────────────────────
df["URL_norm"] = df["URL"].str.replace(r"\.html$", "", regex=True)

df_consolidated = (
    df.groupby("URL_norm", as_index=False)["Visits"]
    .sum()
    .rename(columns={"URL_norm": "URL"})
)

print(f"Total filas después de consolidar: {len(df_consolidated)}")

# Guardar el CSV consolidado
output_csv = "URL_visitas_consolidadas.csv"
df_consolidated.to_csv(output_csv, index=False, encoding="utf-8-sig")
print(f"CSV consolidado guardado: {output_csv}")

# ── 3. Leer el archivo base Excel ─────────────────────────────────────────────
excel_path = "notas_blog_telcel_20260322_143215.xlsx"
df_base = pd.read_excel(excel_path)

print(f"\nArchivo base - columnas: {df_base.columns.tolist()}")
print(f"Archivo base - filas: {len(df_base)}")
print(df_base.head(3).to_string())

# ── 4. Detectar la columna de URL en el archivo base ─────────────────────────
# Buscar columna que contenga URLs de telcel
url_col = None
for col in df_base.columns:
    sample = df_base[col].dropna().astype(str)
    if sample.str.contains("telcel.com", case=False).any():
        url_col = col
        print(f"\nColumna URL detectada: '{url_col}'")
        break

if url_col is None:
    print("\nNo se detectó columna de URL automáticamente.")
    print("Columnas disponibles:", df_base.columns.tolist())
    exit(1)

# ── 5. Normalizar URLs del archivo base para el cruce ────────────────────────
def normalize_url(url):
    if pd.isna(url):
        return ""
    url = str(url).strip()
    # Quitar protocolo si existe
    url = url.replace("https://", "").replace("http://", "")
    # Quitar .html al final
    if url.endswith(".html"):
        url = url[:-5]
    return url

df_base["URL_norm"] = df_base[url_col].apply(normalize_url)
df_consolidated["URL_norm"] = df_consolidated["URL"].apply(normalize_url)

# ── 6. Hacer el cruce (merge) ─────────────────────────────────────────────────
df_merged = df_base.merge(
    df_consolidated[["URL_norm", "Visits"]],
    on="URL_norm",
    how="left"
)

df_merged = df_merged.rename(columns={"Visits": "visitas"})
df_merged = df_merged.drop(columns=["URL_norm"])

coincidencias = df_merged["visitas"].notna().sum()
print(f"\nCoincidencias encontradas: {coincidencias} de {len(df_base)} filas")

# ── 7. Guardar el archivo base con la columna de visitas ─────────────────────
output_excel = "notas_blog_telcel_con_visitas.xlsx"
df_merged.to_excel(output_excel, index=False)
print(f"Archivo base con visitas guardado: {output_excel}")

# Mostrar ejemplos de coincidencias
print("\nEjemplos de coincidencias:")
print(df_merged[df_merged["visitas"].notna()][[url_col, "visitas"]].head(10).to_string())

# ── 8. Aplicar colores en columna Título para notas anteriores al 01/01/2025 ──
FILL_ROJO    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_VERDE   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

FECHA_CORTE = pd.Timestamp("2025-01-01")

wb = load_workbook(output_excel)
ws = wb.active

# Encontrar índices de columnas por nombre (fila 1 es el header)
header = {cell.value: cell.column for cell in ws[1]}
col_titulo  = header.get("Título") or header.get("Titulo")
col_fecha   = header.get("Fecha")
col_visitas = header.get("visitas")

if not all([col_titulo, col_fecha, col_visitas]):
    print(f"Columnas en header: {list(header.keys())}")
    raise ValueError("No se encontraron las columnas necesarias en el Excel.")

conteo = {"rojo": 0, "amarillo": 0, "verde": 0, "omitidas": 0}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    fecha_cell   = row[col_fecha - 1]
    titulo_cell  = row[col_titulo - 1]
    visitas_cell = row[col_visitas - 1]

    # Parsear fecha
    fecha_val = fecha_cell.value
    if fecha_val is None:
        conteo["omitidas"] += 1
        continue

    if isinstance(fecha_val, str):
        try:
            fecha_val = pd.to_datetime(fecha_val, dayfirst=True)
        except Exception:
            conteo["omitidas"] += 1
            continue

    # Solo notas anteriores al 01/01/2025
    if pd.Timestamp(fecha_val) >= FECHA_CORTE:
        conteo["omitidas"] += 1
        continue

    visitas = visitas_cell.value
    if visitas is None or visitas < 400:
        titulo_cell.fill = FILL_ROJO
        conteo["rojo"] += 1
    elif visitas < 1000:
        titulo_cell.fill = FILL_AMARILLO
        conteo["amarillo"] += 1
    else:
        titulo_cell.fill = FILL_VERDE
        conteo["verde"] += 1

wb.save(output_excel)
print(f"\nColores aplicados (notas anteriores al 01/01/2025):")
print(f"  Rojo    (< 400 o sin visitas): {conteo['rojo']}")
print(f"  Amarillo (400–999 visitas):   {conteo['amarillo']}")
print(f"  Verde   (>= 300 visitas):      {conteo['verde']}")
print(f"  Omitidas (fecha >= 2025):      {conteo['omitidas']}")
